"""
Генератор выходного файла конфигурации для PProg.
Преобразует внутреннее представление конфигурации в текстовый формат,
совместимый с импортом в PProg.
"""

from pathlib import Path
from typing import Optional
from datetime import datetime

from data.models import Configuration, Device, Partition, Relay, RelayProgram


class PProgExporter:
    """Экспортер конфигурации в формат PProg."""
    
    def __init__(self, configuration: Configuration):
        self.configuration = configuration
    
    def generate_txt(self, output_path: str | Path) -> bool:
        """
        Генерация текстового файла конфигурации.
        
        Args:
            output_path: Путь к выходному файлу
            
        Returns:
            True если успешно, False иначе
        """
        try:
            output_path = Path(output_path)
            
            lines = []
            
            # Заголовок файла
            lines.append(f"; Конфигурация PProg")
            lines.append(f"; Проект: {self.configuration.project_name}")
            lines.append(f"; Дата генерации: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            lines.append(f"; Инструмент: Project-to-PProg")
            lines.append("")
            
            # Секция устройств
            lines.append("[DEVICES]")
            lines.append("; Address, Type, Description, Version")
            for device in self.configuration.devices:
                version_str = device.version if device.version else ""
                lines.append(f"{device.address}, {device.device_type}, \"{device.description}\", {version_str}")
            lines.append("")
            
            # Секция разделов
            lines.append("[PARTITIONS]")
            lines.append("; PartitionID, Name, Enabled")
            for partition in self.configuration.partitions:
                enabled_str = "1" if partition.enabled else "0"
                lines.append(f"{partition.partition_id}, \"{partition.name}\", {enabled_str}")
                
                # Зоны в разделе
                if partition.zones:
                    lines.append(f";   Zones for Partition {partition.partition_id}")
                    for zone in partition.zones:
                        zone_type_str = zone.zone_type.value if hasattr(zone.zone_type, 'value') else str(zone.zone_type)
                        lines.append(f";   Zone {zone.zone_number}: Type={zone_type_str}, Addr={zone.address}, Algo={zone.algorithm}")
            lines.append("")
            
            # Секция реле
            lines.append("[RELAYS]")
            lines.append("; DeviceAddr, RelayNum, Program, Delay, ActivationTime, Description")
            for relay in self.configuration.relays:
                program_value = relay.program.value if hasattr(relay.program, 'value') else int(relay.program)
                partitions_str = ",".join(map(str, relay.partitions)) if relay.partitions else ""
                lines.append(
                    f"{relay.device_address}, {relay.relay_number}, {program_value}, "
                    f"{relay.delay}, {relay.activation_time}, \"{relay.description}\""
                )
                if relay.partitions:
                    lines.append(f";   Controlled by partitions: {partitions_str}")
            lines.append("")
            
            # Секция сценариев управления
            if self.configuration.scenarios:
                lines.append("[MANAGEMENT_SCENARIOS]")
                lines.append("; ScenarioID, Name, Enabled")
                for scenario in self.configuration.scenarios:
                    enabled_str = "1" if scenario.enabled else "0"
                    lines.append(f"{scenario.scenario_id}, \"{scenario.name}\", {enabled_str}")
                    
                    if scenario.conditions:
                        lines.append(f";   Conditions:")
                        for cond in scenario.conditions:
                            lines.append(f";   - {cond}")
                    
                    if scenario.actions:
                        lines.append(f";   Actions:")
                        for action in scenario.actions:
                            lines.append(f";   - {action}")
                lines.append("")
            
            # Секция валидации
            lines.append("[VALIDATION]")
            errors = self.configuration.validate()
            if errors:
                lines.append("; ERRORS FOUND:")
                for error in errors:
                    lines.append(f"; ERROR: {error}")
            else:
                lines.append("; No validation errors found")
            lines.append("")
            
            # Запись в файл
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write('\n'.join(lines))
            
            return True
            
        except Exception as e:
            print(f"Ошибка при генерации файла: {e}")
            return False
    
    def generate_json(self, output_path: str | Path) -> bool:
        """
        Генерация JSON файла конфигурации (альтернативный формат).
        
        Args:
            output_path: Путь к выходному файлу
            
        Returns:
            True если успешно, False иначе
        """
        try:
            import json
            output_path = Path(output_path)
            
            data = {
                "project_name": self.configuration.project_name,
                "generated_at": datetime.now().isoformat(),
                "devices": [
                    {
                        "address": d.address,
                        "type": d.device_type,
                        "description": d.description,
                        "version": d.version,
                        "status": d.status.value
                    }
                    for d in self.configuration.devices
                ],
                "partitions": [
                    {
                        "id": p.partition_id,
                        "name": p.name,
                        "enabled": p.enabled,
                        "zones": [
                            {
                                "number": z.zone_number,
                                "type": z.zone_type.value if hasattr(z.zone_type, 'value') else str(z.zone_type),
                                "address": z.address,
                                "algorithm": z.algorithm,
                                "enabled": z.enabled
                            }
                            for z in p.zones
                        ]
                    }
                    for p in self.configuration.partitions
                ],
                "relays": [
                    {
                        "device_address": r.device_address,
                        "relay_number": r.relay_number,
                        "program": r.program.value if hasattr(r.program, 'value') else int(r.program),
                        "partitions": r.partitions,
                        "delay": r.delay,
                        "activation_time": r.activation_time,
                        "description": r.description
                    }
                    for r in self.configuration.relays
                ],
                "scenarios": [
                    {
                        "id": s.scenario_id,
                        "name": s.name,
                        "enabled": s.enabled,
                        "conditions": s.conditions,
                        "actions": s.actions
                    }
                    for s in self.configuration.scenarios
                ],
                "validation_errors": self.configuration.validate()
            }
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            
            return True
            
        except Exception as e:
            print(f"Ошибка при генерации JSON: {e}")
            return False
    
    def generate_excel_report(self, output_path: str | Path) -> bool:
        """
        Генерация отчета в формате Excel.
        
        Args:
            output_path: Путь к выходному файлу
            
        Returns:
            True если успешно, False иначе
        """
        try:
            # Проверяем доступность openpyxl
            try:
                from openpyxl import Workbook
            except ImportError:
                print("openpyxl не установлен. Установите: pip install openpyxl")
                return False
            
            output_path = Path(output_path)
            wb = Workbook()
            
            # Удаляем стандартный лист
            wb.remove(wb.active)
            
            # Лист устройств
            ws_devices = wb.create_sheet("Devices")
            ws_devices.append(["Address", "Type", "Description", "Version", "Status"])
            for device in self.configuration.devices:
                ws_devices.append([
                    device.address,
                    device.device_type,
                    device.description,
                    device.version or "",
                    device.status.value
                ])
            
            # Лист разделов
            ws_partitions = wb.create_sheet("Partitions")
            ws_partitions.append(["Partition ID", "Name", "Enabled", "Zone Count"])
            for partition in self.configuration.partitions:
                ws_partitions.append([
                    partition.partition_id,
                    partition.name,
                    "Yes" if partition.enabled else "No",
                    len(partition.zones)
                ])
            
            # Лист зон
            ws_zones = wb.create_sheet("Zones")
            ws_zones.append(["Partition", "Zone Number", "Type", "Address", "Algorithm", "Enabled"])
            for partition in self.configuration.partitions:
                for zone in partition.zones:
                    ws_zones.append([
                        partition.partition_id,
                        zone.zone_number,
                        zone.zone_type.value if hasattr(zone.zone_type, 'value') else str(zone.zone_type),
                        zone.address,
                        zone.algorithm,
                        "Yes" if zone.enabled else "No"
                    ])
            
            # Лист реле
            ws_relays = wb.create_sheet("Relays")
            ws_relays.append(["Device Address", "Relay Num", "Program", "Delay", "Activation Time", "Description", "Partitions"])
            for relay in self.configuration.relays:
                ws_relays.append([
                    relay.device_address,
                    relay.relay_number,
                    relay.program.name if hasattr(relay.program, 'name') else str(relay.program),
                    relay.delay,
                    relay.activation_time,
                    relay.description,
                    ", ".join(map(str, relay.partitions)) if relay.partitions else ""
                ])
            
            # Лист сценариев
            if self.configuration.scenarios:
                ws_scenarios = wb.create_sheet("Scenarios")
                ws_scenarios.append(["ID", "Name", "Enabled", "Conditions", "Actions"])
                for scenario in self.configuration.scenarios:
                    conditions_str = "; ".join([str(c) for c in scenario.conditions])
                    actions_str = "; ".join([str(a) for a in scenario.actions])
                    ws_scenarios.append([
                        scenario.scenario_id,
                        scenario.name,
                        "Yes" if scenario.enabled else "No",
                        conditions_str,
                        actions_str
                    ])
            
            # Лист валидации
            ws_validation = wb.create_sheet("Validation")
            ws_validation.append(["Check", "Result"])
            errors = self.configuration.validate()
            if errors:
                for error in errors:
                    ws_validation.append(["ERROR", error])
            else:
                ws_validation.append(["STATUS", "No errors found"])
            
            wb.save(output_path)
            return True
            
        except Exception as e:
            print(f"Ошибка при генерации Excel: {e}")
            return False


def export_configuration(
    configuration: Configuration,
    output_path: str | Path,
    format: str = "txt"
) -> bool:
    """
    Удобная функция для экспорта конфигурации.
    
    Args:
        configuration: Конфигурация для экспорта
        output_path: Путь к выходному файлу
        format: Формат экспорта ("txt", "json", "excel")
        
    Returns:
        True если успешно, False иначе
    """
    exporter = PProgExporter(configuration)
    
    if format.lower() == "txt":
        return exporter.generate_txt(output_path)
    elif format.lower() == "json":
        return exporter.generate_json(output_path)
    elif format.lower() in ["excel", "xlsx"]:
        return exporter.generate_excel_report(output_path)
    else:
        print(f"Неподдерживаемый формат: {format}")
        return False
